import os
os.system("pip install termcolor")
os.system("cls")
from termcolor import colored
try:
	from openpyxl import load_workbook ,Workbook
	from openpyxl.styles import Font
	from tqdm import trange
	from PIL import Image, ImageFont, ImageDraw
	
except ModuleNotFoundError:
	print(colored("Requied Modules Not Found \n Installing .....","red"))
	os.system("python -m pip install -r Requirement.txt")
	print(colored("Succesfully Installed \nPlease Relaunch Script","green"))
	os.system("python openpyxl.py")
	exit()
def PrintCert(name,merit):
	if(merit==False):
		my_img = Image.open("Cert.PNG")

		font_i=ImageFont.truetype('tango.ttf',80)

		img_edit=ImageDraw.ImageDraw(my_img)
		img_edit.text((240,250),name,(0, 0,0),font=font_i)
		
		# path_save=os.getcwd

		rename=(os.getcwd()+"/s_cert/"+name+".PNG")
		my_img.save(rename)
	else:
		my_img = Image.open("ach.jpg")

		font_i=ImageFont.truetype('tango.ttf',80)

		img_edit=ImageDraw.ImageDraw(my_img)
		img_edit.text((240,250),name,(0, 0,0),font=font_i)
	
		# path_save=os.getcwd

		rename=(os.getcwd()+"/s_cert/"+name+".PNG")
		my_img.save(rename)
path = 'stud.xlsx'
wb=load_workbook(filename=path)
ws = wb.active
# m_row = len(ws['a'])
m_row=2
for i in trange(2,m_row+1,desc=f"Printing Certificates",unit="mb"):
	row1=ws.cell(row=i,column=1)
	row_merit=ws.cell(row=i,column=5)
	names = row1.value
	if(row_merit.value=="Institution - Merit"):
		merit=True
	else:
		merit=False
	PrintCert(names,merit)
	
wb.close()

input()



